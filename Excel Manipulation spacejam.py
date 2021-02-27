import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
#wb=Workbook()
#No longer required after creation of excel sheet
wb = load_workbook('DataRepo.xlsx')
ws= wb.active
#Opens a sheet to start working on
ws.title="Input Data"
wb.save('DataRepo.xlsx')
flag=0
global count
count=1
#Format of Data Storage: Name, Severity, Location
class user:
    global N,D,S,L
    N= str()
    D= datetime.datetime(2020,1,1)
    #YY,MM,DD
    S= str()
    L= str()
    def input(self):
        global N,D,S,L
        N=input("Enter your name")
        D=input("Date tested")
        S=input("Calculated severity @Anish")
        L=input("Vector location @abhay")
global U
U=user()
U.input()#This inputs
#Function to write
def write():
    SN=ws.cell(row=count,column=1)
    SN.value=N
    SD=ws.cell(row=count,column=2)
    SD.value=D
    SS=ws.cell(row=count,column=3)
    SS.value=S
    SL=ws.cell(row=count,column=4)
    SL.value=L
    wb.save('DataRepo.xlsx')
write()
#Works till here

#Function to read
def read():
    SN=ws.cell(row=count,column=1)
    N=SN.value
    SD=ws.cell(row=count,column=2)
    D=SD.value
    SS=ws.cell(row=count,column=3)
    S=SS.value
    SL=ws.cell(row=count,column=4)
    L=SL.value

#function usage

write()
read()
