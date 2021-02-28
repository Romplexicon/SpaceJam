
from datetime import date
from typing import Dict, Any

from openpyxl import Workbook, load_workbook
print("~~~~~~~~~~~~~~~~~~~~~~~~COVID RISK CALCULATOR~~~~~~~~~~~~~~~~~~~~~~")
# wb=Workbook()
# No longer required after creation of excel sheet
wb = load_workbook('/Users/Anish/Downloads/DataRepo.xlsx')
ws = wb.active
# Opens a sheet to start working on
ws.title = "Input Data"
wb.save('/Users/Anish/Downloads/DataRepo.xlsx')
flag = 0
global count
count = 1
global count_positive
count_positive = 0
global High_Risk_Sections
High_Risk_Sections = []
today = date.today()
present = today.strftime("%Y/%m/%d")
print(f"Today's date is {present}")
n = int(input("Enter the number of students in the survey "))


class severity():
    d1 = today.strftime("%Y/%m/%d")
    d = d1.split("/")
    status = ""
    D2 = input("Please enter the date you last tested negative for Covid-19. Enter in the form of YY/MM/DD : ")
    d2 = D2.split("/")
    f_date = date(int(d[0]), int(d[1]), int(d[2]))
    l_date = date(int(d2[0]), int(d2[1]), int(d2[2]))
    delta = f_date - l_date
    x = delta.days
    if x >= 90:
        status = "High Risk"
    else:
        status = "Low Risk"
    # Format of Data Storage: Name, Severity, Location
class user:
    global N, D, S, L, C, n
    N = str()
    S = str()
    L = str()
    C = str()
    def user_input(self):
        global N, D, S, L
        N = input("Enter your name : ")
        C = input("Enter your section in Block letters: ")
        D = a.D2
        S = a.status
        print(S)
        if S == "High Risk":
            High_Risk_Sections.append(C.upper())
global U

    # Function to write
def write():
    SN = ws.cell(row=count, column=1)
    SN.value = N
    SD = ws.cell(row=count, column=2)
    SD.value = D
    SS = ws.cell(row=count, column=3)
    SS.value = S
    SL = ws.cell(row=count, column=4)
    SL.value = L
    SC = ws.cell(row=count, column=5)
    SC.value = C
    wb.save('/Users/Anish/Downloads/DataRepo.xlsx')


    # Function to read
def read():
    SN = ws.cell(row=count, column=1)
    N = SN.value
    SD = ws.cell(row=count, column=2)
    D = SD.value
    SS = ws.cell(row=count, column=3)
    S = SS.value
    SL = ws.cell(row=count, column=4)
    L = SL.value
    SC = ws.cell(row=count, column=5)
    C = SC.value
la = []
lb = []
lc = []
d = {}
    # function usage
for i in range(1, n+1):
    a = severity()
    U = user()
    U.user_input()  # This inputs
    write()
    read()
    count += 1
    x = ws.cell(row = i, column = 5).value
    if x == "A" :
       la.append(ws.cell(row = i, column = 1).value)
       d['A'] = la
    elif x.upper == "B":
        lb.append(ws.cell(row = i, column = 1).value)
        d['B'] = lb
    elif x.upper == "C":
        lc.append(ws.cell(row = i, column = 1).value)
        d['C'] = lc
    else:
        print("Sorry! This version of the app is only taking for the first three sections at this point in time.")
        break

print(d)
print(f"These are the sections which are at a high risk : {High_Risk_Sections}")
print("If you are in a high risk section, please actively stay away from others.")
